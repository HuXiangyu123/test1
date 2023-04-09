#读写文本文件
# file=open('test.txt','w')#绝对路径也行
# filejuedui=open(r'C:\Users\lenovo\Desktop\test.txt','w')#绝对路径也行,前面加个r
# data=file.read()
# file.close()
# print(data)

# 读写图片等
# filepng=open('test.png','rb')
# data2=filepng.read()
# filepng.close()

#readline()方法
# file_obj=open('test.txt','rt',encoding='utf-8')
# content=file_obj.readline()
# print(content)
# file_obj.close()

#with 方法
# import requests
# with open('infos/user1.csv',mode='r',encoding='utf-8') as file_object:
#     file_object.readline()
#     for line in file_object:
#         user_id,username,url=line.strip().split(',') #strip()去掉空格
#         print(user_id,username)
#
#         res=requests.get(url)(
#             url=url,
#             headers={
#                 "User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.132 Safari/537.36"
#             }
#         )
#         with open('infos/{}.jpg'.format(username),'wb') as file_objectpic:
#             file_objectpic.write(res.content)

from openpyxl import load_workbook
wb=load_workbook('infos/test.xlsx')

sheet=wb['Sheet1']

print(wb.sheetnames)
cell=sheet.cell(1,1)
print(cell.value)
